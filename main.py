# -*- coding: utf-8 -*-
"""
EXPRESS CONSUMO OC — By Maicon
Versão Android profissional com armazenamento seguro, leitura de PDF e relatórios premium.
"""

import os
import re
import sys
import math
import shutil
import unicodedata
import traceback
from pathlib import Path
from datetime import datetime

from kivy.app import App
from kivy.clock import Clock
from kivy.lang import Builder
from kivy.metrics import dp
from kivy.properties import StringProperty, BooleanProperty
from kivy.uix.boxlayout import BoxLayout
from kivy.utils import platform

KV = r'''
#:import dp kivy.metrics.dp

<PremiumButton@Button>:
    size_hint_y: None
    height: dp(58)
    font_size: "17sp"
    bold: True
    color: 1, 1, 1, 1
    background_normal: ""
    background_down: ""
    background_color: 0, 0, 0, 0
    canvas.before:
        Color:
            rgba: 0.02, 0.23, 0.56, 1
        RoundedRectangle:
            pos: self.pos
            size: self.size
            radius: [dp(16),]
        Color:
            rgba: 0.12, 0.53, 0.95, 0.35
        Line:
            rounded_rectangle: self.x, self.y, self.width, self.height, dp(16)
            width: 1.15

<RootUI>:
    orientation: "vertical"
    padding: dp(14)
    spacing: dp(10)
    canvas.before:
        Color:
            rgba: 0.018, 0.030, 0.070, 1
        Rectangle:
            pos: self.pos
            size: self.size
        Color:
            rgba: 0.06, 0.12, 0.23, 0.92
        RoundedRectangle:
            pos: self.x + dp(8), self.top - dp(174)
            size: self.width - dp(16), dp(160)
            radius: [dp(24),]
        Color:
            rgba: 0.95, 0.66, 0.15, 0.85
        Line:
            points: self.x + dp(28), self.top - dp(166), self.right - dp(28), self.top - dp(166)
            width: 1.2

    BoxLayout:
        orientation: "vertical"
        size_hint_y: None
        height: dp(164)
        padding: dp(8), dp(10), dp(8), dp(8)
        spacing: dp(4)

        Label:
            text: "EXPRESS CONSUMO OC"
            font_size: "29sp"
            bold: True
            color: 1, 1, 1, 1
            halign: "center"
            valign: "middle"
            text_size: self.size

        Label:
            text: "Cálculo de Consumo • Conferência de Ordem de Compra"
            font_size: "15sp"
            bold: True
            color: 0.95, 0.67, 0.18, 1
            halign: "center"
            valign: "middle"
            text_size: self.size

        Label:
            text: "Relatórios profissionais • PDF e Excel • Armazenamento seguro"
            font_size: "12sp"
            color: 0.76, 0.82, 0.92, 1
            halign: "center"
            valign: "middle"
            text_size: self.size

        BoxLayout:
            size_hint_y: None
            height: dp(44)
            spacing: dp(12)
            TextInput:
                id: data_ini
                text: root.data_inicio
                hint_text: "Início"
                multiline: False
                font_size: "18sp"
                halign: "center"
                foreground_color: 1,1,1,1
                background_color: 0.03,0.07,0.14,0.92
                cursor_color: 0.95,0.65,0.13,1
            TextInput:
                id: data_fim
                text: root.data_fim
                hint_text: "Fim"
                multiline: False
                font_size: "18sp"
                halign: "center"
                foreground_color: 1,1,1,1
                background_color: 0.03,0.07,0.14,0.92
                cursor_color: 0.95,0.65,0.13,1

    PremiumButton:
        text: "Selecionar PDF Analítico"
        on_release: root.escolher_analitico()

    PremiumButton:
        text: "Selecionar PDF Ordem de Compra"
        on_release: root.escolher_oc()

    PremiumButton:
        text: "Pasta automática segura"
        on_release: root.definir_pasta_automatica()

    PremiumButton:
        text: "Criar consumo semanal"
        on_release: root.gerar_consumo()

    PremiumButton:
        text: "Conferir Ordem de Compra"
        on_release: root.conferir_oc()

    BoxLayout:
        orientation: "vertical"
        padding: dp(12)
        spacing: dp(4)
        canvas.before:
            Color:
                rgba: 0.035, 0.055, 0.105, 0.96
            RoundedRectangle:
                pos: self.pos
                size: self.size
                radius: [dp(18),]
            Color:
                rgba: 0.95, 0.67, 0.18, 0.18
            Line:
                rounded_rectangle: self.x, self.y, self.width, self.height, dp(18)
                width: 1
        Label:
            text: "STATUS OPERACIONAL"
            size_hint_y: None
            height: dp(24)
            font_size: "12sp"
            bold: True
            color: 0.95, 0.67, 0.18, 1
            halign: "left"
            valign: "middle"
            text_size: self.size
        ScrollView:
            do_scroll_x: False
            Label:
                id: status
                text: root.status
                size_hint_y: None
                height: max(self.texture_size[1] + dp(30), dp(180))
                font_size: "13sp"
                color: 0.90, 0.93, 0.98, 1
                halign: "left"
                valign: "top"
                text_size: self.width, None

    Label:
        text: root.rodape
        size_hint_y: None
        height: dp(26)
        font_size: "13sp"
        color: 0.72, 0.76, 0.84, 1
        halign: "right"
        valign: "middle"
        text_size: self.size
'''

Builder.load_string(KV)


def hoje_br() -> str:
    return datetime.now().strftime("%d/%m/%Y")


def normalizar_nome(txt: str) -> str:
    if not txt:
        return ""
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    txt = txt.upper().strip()
    txt = re.sub(r"[^A-Z0-9 KGUNLT/.-]+", " ", txt)
    palavras_remover = {
        "KG", "UN", "UND", "LT", "L", "CX", "PCT", "PACOTE", "CAIXA",
        "CONGELADO", "RESFRIADO", "NAO", "PERECIVEL", "PERECIVEIS"
    }
    partes = [p for p in txt.split() if p not in palavras_remover]
    return " ".join(partes)


def parse_qtd(valor: str) -> float:
    if valor is None:
        return 0.0
    s = str(valor).strip()
    s = re.sub(r"[^0-9,.-]", "", s)
    if not s:
        return 0.0
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def fmt_qtd(v: float) -> str:
    try:
        if abs(v - round(v)) < 0.001:
            return str(int(round(v)))
        return f"{v:.3f}".replace(".", ",").rstrip("0").rstrip(",")
    except Exception:
        return str(v)


def extrair_texto_pdf(caminho_pdf: str) -> str:
    """Extrai texto do PDF usando pypdf, biblioteca mais leve e compatível com APK Android.
    Evita pdfplumber/pandas/reportlab, que costumam quebrar a compilação no Buildozer.
    """
    texto = ""
    try:
        from pypdf import PdfReader
        reader = PdfReader(caminho_pdf)
        for page in reader.pages:
            try:
                texto += (page.extract_text() or "") + "\n"
            except Exception:
                texto += "\n"
    except Exception as e:
        raise RuntimeError(f"Falha ao ler PDF com pypdf: {e}")
    return texto


def extrair_analitico(caminho_pdf: str, data_ini: str, data_fim: str):
    """Extrator tolerante para relatório analítico: produto, unidade, quantidade e data."""
    texto = extrair_texto_pdf(caminho_pdf)
    linhas = [l.strip() for l in texto.splitlines() if l.strip()]
    itens = []
    data_atual = ""
    servico_atual = ""

    for linha in linhas:
        m_data = re.search(r"(?:Data|DATA)\s*[:\-]?\s*(\d{2}/\d{2}/\d{4})", linha)
        if m_data:
            data_atual = m_data.group(1)

        m_serv = re.search(r"\b(ALMO[CÇ]O|JANTAR|JANTA|CEIA)\b", linha, re.I)
        if m_serv:
            servico_atual = m_serv.group(1).upper().replace("Ç", "C")

        # Padrão comum: código + descrição + qtd + unidade
        m = re.search(
            r"(?P<cod>\d+\.\d+\.\d+(?:\.\d+)*)\s+(?P<nome>.+?)\s+(?P<qtd>\d{1,6}(?:[\.,]\d{1,3})?)\s+(?P<un>KG|UN|UND|LT|L|CX|PCT)\b",
            linha,
            re.I,
        )
        if m:
            nome = m.group("nome").strip(" -")
            qtd = parse_qtd(m.group("qtd"))
            un = m.group("un").upper().replace("UND", "UN").replace("L", "LT")
            if nome and qtd > 0:
                itens.append({
                    "codigo": m.group("cod"),
                    "produto": nome.upper(),
                    "chave": normalizar_nome(nome),
                    "unidade": un,
                    "necessario": qtd,
                    "data": data_atual,
                    "turno": servico_atual,
                    "origem": "ANALITICO",
                })

    if not itens:
        # fallback: linhas com nome e quantidade/unidade sem código
        for linha in linhas:
            m = re.search(r"(?P<nome>[A-Za-zÁ-Úá-ú0-9 /.-]{4,})\s+(?P<qtd>\d{1,6}(?:[\.,]\d{1,3})?)\s+(?P<un>KG|UN|UND|LT|L|CX|PCT)\b", linha, re.I)
            if m:
                nome = m.group("nome").strip(" -")
                qtd = parse_qtd(m.group("qtd"))
                if qtd > 0:
                    itens.append({
                        "codigo": "",
                        "produto": nome.upper(),
                        "chave": normalizar_nome(nome),
                        "unidade": m.group("un").upper().replace("UND", "UN").replace("L", "LT"),
                        "necessario": qtd,
                        "data": data_atual,
                        "turno": servico_atual,
                        "origem": "ANALITICO",
                    })

    return consolidar_itens(itens, campo_qtd="necessario")


def extrair_oc(caminho_pdf: str):
    """Extrator tolerante para Ordem de Compra: produto, unidade, quantidade, entrega/utilização."""
    texto = extrair_texto_pdf(caminho_pdf)
    linhas = [l.strip() for l in texto.splitlines() if l.strip()]
    itens = []

    for linha in linhas:
        # Modelo esperado: Produto UN Quantidade Entrega Utilização
        m = re.search(
            r"(?P<nome>[A-Za-zÁ-Úá-ú0-9 /().,-]{4,}?)\s+(?P<un>KG|UN|UND|LT|L|CX|PCT)\s+(?P<qtd>\d{1,6}(?:[\.,]\d{1,3})?)\s*(?P<resto>.*)$",
            linha,
            re.I,
        )
        if m:
            nome = m.group("nome").strip(" -")
            if any(x in nome.upper() for x in ["PRODUTO", "QUANTIDADE", "SOLICITACAO", "FILIAL"]):
                continue
            qtd = parse_qtd(m.group("qtd"))
            if qtd <= 0:
                continue
            datas = re.findall(r"\d{2}/\d{2}/\d{4}", m.group("resto"))
            itens.append({
                "produto": nome.upper(),
                "chave": normalizar_nome(nome),
                "unidade": m.group("un").upper().replace("UND", "UN").replace("L", "LT"),
                "comprado": qtd,
                "entrega": datas[0] if len(datas) >= 1 else "",
                "utilizacao": datas[-1] if datas else "",
                "origem": "OC",
            })

    return consolidar_itens(itens, campo_qtd="comprado")


def consolidar_itens(itens, campo_qtd: str):
    consol = {}
    for it in itens:
        chave = f"{it.get('chave','')}|{it.get('unidade','')}|{it.get('data') or it.get('utilizacao') or ''}"
        if chave not in consol:
            consol[chave] = dict(it)
        else:
            consol[chave][campo_qtd] = consol[chave].get(campo_qtd, 0) + it.get(campo_qtd, 0)
            # mantém nome mais comprido, geralmente mais descritivo
            if len(it.get("produto", "")) > len(consol[chave].get("produto", "")):
                consol[chave]["produto"] = it.get("produto", "")
    return list(consol.values())


def comparar_consumo_oc(consumo, oc):
    oc_por_chave = {}
    for item in oc:
        base = item["chave"]
        oc_por_chave.setdefault(base, []).append(item)

    faltantes = []
    sobras = []
    conferidos = []
    usados_oc = set()

    for nec in consumo:
        chave = nec["chave"]
        candidatos = oc_por_chave.get(chave, [])
        # busca aproximada simples
        if not candidatos:
            for k, vals in oc_por_chave.items():
                if chave and (chave in k or k in chave):
                    candidatos = vals
                    break
        comprado = sum(c.get("comprado", 0) for c in candidatos)
        for c in candidatos:
            usados_oc.add(id(c))
        necessario = nec.get("necessario", 0)
        diff = comprado - necessario
        linha = {
            "produto": nec.get("produto", ""),
            "unidade": nec.get("unidade", ""),
            "necessario": necessario,
            "comprado": comprado,
            "data": nec.get("data", ""),
            "observacao": "OK" if diff >= -0.001 else "FALTA CRÍTICA PARA DATA DE CONSUMO",
        }
        conferidos.append(linha)
        if diff < -0.001:
            linha["faltante"] = abs(diff)
            faltantes.append(linha)
        elif diff > 0.001:
            linha["sobra"] = diff
            sobras.append(linha)

    # itens comprados sem consumo identificado entram como sobra/excedente
    for item in oc:
        if id(item) not in usados_oc:
            sobras.append({
                "produto": item.get("produto", ""),
                "unidade": item.get("unidade", ""),
                "necessario": 0,
                "comprado": item.get("comprado", 0),
                "sobra": item.get("comprado", 0),
                "data": item.get("utilizacao", ""),
                "observacao": "ITEM NA OC SEM CONSUMO IDENTIFICADO NO ANALÍTICO",
            })
    return faltantes, sobras, conferidos


def _pdf_escape(texto):
    return str(texto).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)").replace("\r", " ").replace("\n", " ")


def _quebrar_linha(texto, max_chars=110):
    texto = str(texto)
    partes = []
    atual = ""
    for palavra in texto.split():
        if len(atual) + len(palavra) + 1 > max_chars:
            if atual:
                partes.append(atual)
            atual = palavra
        else:
            atual = (atual + " " + palavra).strip()
    if atual:
        partes.append(atual)
    return partes or [""]



def _pdf_color(r, g, b, stroke=False):
    op = "RG" if stroke else "rg"
    return f"{r:.3f} {g:.3f} {b:.3f} {op}\n"


def _pdf_rect(x, y, w, h, fill=True):
    return f"{x:.2f} {y:.2f} {w:.2f} {h:.2f} re {'f' if fill else 'S'}\n"


def _pdf_line(x1, y1, x2, y2, width=0.6):
    return f"{width:.2f} w {x1:.2f} {y1:.2f} m {x2:.2f} {y2:.2f} l S\n"


def _pdf_text(txt, x, y, size=8, bold=False, r=0.05, g=0.07, b=0.12):
    fonte = "/F2" if bold else "/F1"
    safe = _pdf_escape(txt)
    return _pdf_color(r, g, b) + f"BT {fonte} {size} Tf {x:.2f} {y:.2f} Td ({safe}) Tj ET\n"


def _truncate(txt, max_chars):
    txt = str(txt or "")
    return txt if len(txt) <= max_chars else txt[:max(0, max_chars-1)] + "…"


def gerar_pdf_tabela(titulo, subtitulo, linhas, colunas, caminho):
    """Cria PDF profissional sem dependências nativas pesadas, compatível com APK Android."""
    largura = 842
    altura = 595
    margem = 28
    header_h = 78
    footer_h = 30
    table_top = altura - header_h - 18
    row_h = 20
    head_h = 22

    # Larguras planejadas para relatórios operacionais em A4 paisagem.
    default_widths = {
        "PRODUTO": 230,
        "UN": 35,
        "NECESSÁRIO": 72,
        "COMPRADO": 72,
        "FALTANTE": 72,
        "SOBRA": 70,
        "DATA": 70,
        "TURNO": 55,
        "OBSERVAÇÃO": 190,
    }
    usable = largura - margem * 2
    widths = [default_widths.get(c, 95) for c in colunas]
    total = sum(widths)
    if total > usable:
        scale = usable / total
        widths = [max(30, w * scale) for w in widths]

    # Paginação.
    rows_per_page = int((table_top - footer_h - head_h) // row_h)
    data = linhas or [{colunas[0]: "Nenhum item encontrado."}]
    paginas = [data[i:i+rows_per_page] for i in range(0, len(data), rows_per_page)] or [[]]

    objects = []
    pages_ids = []
    objects.append("<< /Type /Catalog /Pages 2 0 R >>")
    objects.append("<< /Type /Pages /Kids [] /Count 0 >>")
    objects.append("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    objects.append("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>")

    for pi, page_rows in enumerate(paginas, start=1):
        c = ""
        # Fundo e cabeçalho
        c += _pdf_color(0.985, 0.988, 0.994) + _pdf_rect(0, 0, largura, altura)
        c += _pdf_color(0.02, 0.04, 0.10) + _pdf_rect(0, altura-header_h, largura, header_h)
        c += _pdf_color(0.95, 0.66, 0.16) + _pdf_rect(0, altura-header_h, 7, header_h)
        c += _pdf_text("EXPRESS COLORADO", margem, altura-30, 10, True, 0.95, 0.66, 0.16)
        c += _pdf_text(titulo[:90], margem, altura-52, 18, True, 1, 1, 1)
        c += _pdf_text(subtitulo[:130], margem, altura-68, 8.5, False, 0.82, 0.88, 0.96)
        c += _pdf_text(f"Criado em {hoje_br()} - By Maicon", largura-180, altura-30, 8.5, False, 0.88, 0.90, 0.95)
        c += _pdf_color(0.95, 0.66, 0.16) + _pdf_rect(margem, altura-header_h-5, largura-margem*2, 2.2)

        # Cabeçalho da tabela
        y = table_top
        x = margem
        c += _pdf_color(0.07, 0.18, 0.36) + _pdf_rect(margem, y-head_h+2, usable, head_h)
        for col, w in zip(colunas, widths):
            c += _pdf_text(_truncate(str(col).upper(), int(w/4.2)), x+4, y-12, 6.7, True, 1, 1, 1)
            x += w
        c += _pdf_color(0.95, 0.66, 0.16) + _pdf_rect(margem, y-head_h+1, usable, 1.2)
        y -= head_h

        # Linhas
        for idx, row in enumerate(page_rows, start=1+(pi-1)*rows_per_page):
            fill = (0.96, 0.975, 0.995) if idx % 2 == 0 else (1, 1, 1)
            c += _pdf_color(*fill) + _pdf_rect(margem, y-row_h+2, usable, row_h)
            x = margem
            for col, w in zip(colunas, widths):
                val = row.get(col, "")
                val = fmt_qtd(val) if isinstance(val, float) else str(val)
                maxc = max(4, int(w / 4.1))
                bold = col in ("FALTANTE", "SOBRA") and val not in ("", "0", "0,0", "0,00")
                color = (0.58, 0.05, 0.06) if col == "FALTANTE" and bold else (0.05, 0.07, 0.12)
                if col == "SOBRA" and bold:
                    color = (0.02, 0.32, 0.18)
                c += _pdf_text(_truncate(val, maxc), x+4, y-11.5, 6.5, bold, *color)
                x += w
            c += _pdf_color(0.82, 0.86, 0.92, True) + _pdf_line(margem, y-row_h+2, margem+usable, y-row_h+2, 0.25)
            y -= row_h

        # Rodapé
        c += _pdf_color(0.02, 0.04, 0.10) + _pdf_rect(0, 0, largura, footer_h)
        c += _pdf_text(f"Página {pi}/{len(paginas)}", margem, 12, 7.5, False, 0.82, 0.86, 0.94)
        c += _pdf_text(f"{hoje_br()} - By Maicon", largura-135, 12, 7.5, False, 0.82, 0.86, 0.94)

        stream_bytes = c.encode("latin-1", "ignore")
        stream = f"<< /Length {len(stream_bytes)} >>\nstream\n".encode("latin-1") + stream_bytes + b"\nendstream"
        content_id = len(objects) + 1
        objects.append(stream)
        page_id = len(objects) + 1
        pages_ids.append(page_id)
        objects.append(f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {largura} {altura}] /Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> /Contents {content_id} 0 R >>")

    kids = " ".join(f"{pid} 0 R" for pid in pages_ids)
    objects[1] = f"<< /Type /Pages /Kids [{kids}] /Count {len(pages_ids)} >>"

    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for i, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf.extend(f"{i} 0 obj\n".encode("latin-1"))
        if isinstance(obj, bytes):
            pdf.extend(obj)
        else:
            pdf.extend(obj.encode("latin-1", "ignore"))
        pdf.extend(b"\nendobj\n")
    xref = len(pdf)
    pdf.extend(f"xref\n0 {len(objects)+1}\n".encode("latin-1"))
    pdf.extend(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        pdf.extend(f"{off:010d} 00000 n \n".encode("latin-1"))
    pdf.extend(f"trailer\n<< /Size {len(objects)+1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF".encode("latin-1"))

    with open(caminho, "wb") as f:
        f.write(pdf)


def gerar_excel(conferidos, faltantes, sobras, caminho):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="0B1F3A")
    gold_fill = PatternFill("solid", fgColor="F2A51A")
    alt_fill = PatternFill("solid", fgColor="F4F7FB")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="0B1F3A", bold=True, size=14)
    thin = Side(style="thin", color="D7DDE8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def escrever_aba(wb, nome, dados):
        ws = wb.create_sheet(nome)
        ws.append(["EXPRESS COLORADO"])
        ws.append([f"{nome.upper()} - Criado em {hoje_br()} - By Maicon"])
        ws.append([])
        if dados:
            colunas = list(dados[0].keys())
        else:
            colunas = ["produto", "unidade", "necessario", "comprado", "faltante", "sobra", "data", "observacao"]
        ws.append([c.upper() for c in colunas])
        header_row = ws.max_row
        for item in dados:
            ws.append([item.get(c, "") for c in colunas])

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(colunas)))
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, len(colunas)))
        ws["A1"].font = title_font
        ws["A2"].font = Font(color="6B7280", bold=True)
        ws["A1"].fill = gold_fill
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A2"].alignment = Alignment(horizontal="center")

        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for row in ws.iter_rows(min_row=header_row+1):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = alt_fill
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = f"A{header_row+1}"
        ws.auto_filter.ref = ws.dimensions
        for idx, col in enumerate(ws.columns, start=1):
            max_len = 12
            for cell in col:
                max_len = max(max_len, len(str(cell.value or "")))
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 42)

    wb = Workbook()
    wb.remove(wb.active)
    escrever_aba(wb, "Conferencia", conferidos)
    escrever_aba(wb, "Faltantes", faltantes)
    escrever_aba(wb, "Sobras", sobras)
    wb.save(caminho)


class RootUI(BoxLayout):
    status = StringProperty("Pronto. Selecione os PDFs e crie os relatórios profissionais.\n")
    data_inicio = StringProperty("11/06/2026")
    data_fim = StringProperty("17/06/2026")
    rodape = StringProperty(f"{hoje_br()} - By Maicon")
    analitico_path = StringProperty("")
    oc_path = StringProperty("")
    output_dir = StringProperty("")
    busy = BooleanProperty(False)

    def log(self, msg):
        self.status += f"\n{msg}"

    def set_status(self, msg):
        self.status = msg

    def get_output_dir(self):
        if platform == "android":
            try:
                from android.storage import app_storage_path
                base = app_storage_path()
            except Exception:
                base = os.getcwd()
        else:
            base = os.path.join(os.getcwd(), "saida")
        pasta = os.path.join(base, "relatorios")
        os.makedirs(pasta, exist_ok=True)
        self.output_dir = pasta
        return pasta

    def definir_pasta_automatica(self):
        pasta = self.get_output_dir()
        self.set_status(
            "Pasta de saída OK.\n"
            "Esta versão não usa content://tree como caminho comum, evitando o erro Invalid URI.\n\n"
            f"Relatórios serão salvos em:\n{pasta}"
        )

    def escolher_analitico(self):
        self._escolher_pdf("analitico")

    def escolher_oc(self):
        self._escolher_pdf("oc")

    def _escolher_pdf(self, tipo):
        try:
            from plyer import filechooser
            filechooser.open_file(on_selection=lambda selection: self._on_pdf_selection(selection, tipo), filters=[("PDF", "*.pdf")])
        except Exception as e:
            self.log(f"Erro ao abrir seletor de PDF: {e}")

    def _on_pdf_selection(self, selection, tipo):
        if not selection:
            self.log("Nenhum arquivo selecionado.")
            return
        uri = selection[0]
        try:
            caminho = self.copiar_uri_para_cache(uri, tipo)
            if tipo == "analitico":
                self.analitico_path = caminho
                self.log(f"Analítico OK:\n{caminho}")
            else:
                self.oc_path = caminho
                self.log(f"Ordem de Compra OK:\n{caminho}")
        except Exception as e:
            self.log(f"Erro ao preparar arquivo {tipo}: {e}")

    def copiar_uri_para_cache(self, uri, prefixo):
        """Copia content:// para arquivo real no cache interno. Corrige o erro Invalid URI."""
        uri = str(uri)
        cache_dir = os.path.join(self.get_output_dir(), "cache")
        os.makedirs(cache_dir, exist_ok=True)
        destino = os.path.join(cache_dir, f"{prefixo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")

        if uri.startswith("content://") and platform == "android":
            from jnius import autoclass
            PythonActivity = autoclass("org.kivy.android.PythonActivity")
            Uri = autoclass("android.net.Uri")
            activity = PythonActivity.mActivity
            resolver = activity.getContentResolver()
            input_stream = resolver.openInputStream(Uri.parse(uri))
            if input_stream is None:
                raise RuntimeError("Android não liberou leitura deste arquivo.")
            from jnius import jarray
            buffer = jarray("b")(8192)
            with open(destino, "wb") as out:
                while True:
                    n = input_stream.read(buffer)
                    if n == -1 or n == 0:
                        break
                    out.write(bytes((buffer[i] & 0xFF) for i in range(n)))
            input_stream.close()
            return destino

        if uri.startswith("file://"):
            origem = uri.replace("file://", "")
        else:
            origem = uri
        if not os.path.exists(origem):
            raise RuntimeError(f"Arquivo não encontrado: {origem}")
        shutil.copyfile(origem, destino)
        return destino

    def _datas_tela(self):
        di = self.ids.data_ini.text.strip()
        df = self.ids.data_fim.text.strip()
        for d in [di, df]:
            if not re.match(r"^\d{2}/\d{2}/\d{4}$", d):
                raise ValueError("Datas devem estar no formato dd/mm/aaaa.")
        return di, df

    def gerar_consumo(self):
        if self.busy:
            return
        Clock.schedule_once(lambda dt: self._gerar_consumo_exec(), 0.1)

    def _gerar_consumo_exec(self):
        try:
            self.busy = True
            self.set_status("Criando consumo semanal...")
            if not self.analitico_path:
                raise RuntimeError("Escolha primeiro o PDF Analítico.")
            di, df = self._datas_tela()
            itens = extrair_analitico(self.analitico_path, di, df)
            if not itens:
                raise RuntimeError("Nenhum item foi extraído do analítico. Verifique se o PDF tem texto selecionável.")
            pasta = self.get_output_dir()
            linhas = []
            for it in itens:
                linhas.append({
                    "PRODUTO": it.get("produto", ""),
                    "UN": it.get("unidade", ""),
                    "NECESSÁRIO": fmt_qtd(it.get("necessario", 0)),
                    "DATA": it.get("data", ""),
                    "TURNO": it.get("turno", ""),
                    "OBSERVAÇÃO": "CONSUMO PREVISTO DO ANALÍTICO",
                })
            caminho_pdf = os.path.join(pasta, "RELATORIO_CONSUMO_SEMANAL.pdf")
            gerar_pdf_tabela(
                "RELATÓRIO DE CONSUMO SEMANAL",
                f"Período: {di} a {df}",
                linhas,
                ["PRODUTO", "UN", "NECESSÁRIO", "DATA", "TURNO", "OBSERVAÇÃO"],
                caminho_pdf,
            )
            self.set_status(f"Consumo semanal criado com sucesso.\n\nItens extraídos: {len(itens)}\nArquivo criado:\n{caminho_pdf}")
        except Exception as e:
            self.set_status(f"ERRO ao criar consumo semanal:\n{e}\n\nDetalhe técnico:\n{traceback.format_exc()[-1200:]}")
        finally:
            self.busy = False

    def conferir_oc(self):
        if self.busy:
            return
        Clock.schedule_once(lambda dt: self._conferir_oc_exec(), 0.1)

    def _conferir_oc_exec(self):
        try:
            self.busy = True
            self.set_status("Conferindo Ordem de Compra...")
            if not self.analitico_path:
                raise RuntimeError("Escolha primeiro o PDF Analítico.")
            if not self.oc_path:
                raise RuntimeError("Escolha primeiro o PDF da Ordem de Compra.")
            di, df = self._datas_tela()
            consumo = extrair_analitico(self.analitico_path, di, df)
            oc = extrair_oc(self.oc_path)
            if not consumo:
                raise RuntimeError("Nenhum item foi extraído do Analítico.")
            if not oc:
                raise RuntimeError("Nenhum item foi extraído da Ordem de Compra.")
            faltantes, sobras, conferidos = comparar_consumo_oc(consumo, oc)
            pasta = self.get_output_dir()

            faltantes_pdf = os.path.join(pasta, "RELATORIO_ITENS_FALTANTES.pdf")
            sobras_pdf = os.path.join(pasta, "RELATORIO_ITENS_SOBRAS.pdf")
            excel = os.path.join(pasta, "CONFERENCIA_CONSUMO_X_OC.xlsx")

            linhas_f = [{
                "PRODUTO": x["produto"], "UN": x["unidade"],
                "NECESSÁRIO": fmt_qtd(x["necessario"]), "COMPRADO": fmt_qtd(x["comprado"]),
                "FALTANTE": fmt_qtd(x.get("faltante", 0)), "DATA": x.get("data", ""),
                "OBSERVAÇÃO": x.get("observacao", ""),
            } for x in faltantes]
            linhas_s = [{
                "PRODUTO": x["produto"], "UN": x["unidade"],
                "NECESSÁRIO": fmt_qtd(x["necessario"]), "COMPRADO": fmt_qtd(x["comprado"]),
                "SOBRA": fmt_qtd(x.get("sobra", 0)), "DATA": x.get("data", ""),
                "OBSERVAÇÃO": x.get("observacao", ""),
            } for x in sobras]

            if not linhas_f:
                linhas_f = [{"PRODUTO": "SEM FALTANTES", "UN": "", "NECESSÁRIO": "", "COMPRADO": "", "FALTANTE": "", "DATA": "", "OBSERVAÇÃO": "Nenhum item faltante identificado."}]
            if not linhas_s:
                linhas_s = [{"PRODUTO": "SEM SOBRAS", "UN": "", "NECESSÁRIO": "", "COMPRADO": "", "SOBRA": "", "DATA": "", "OBSERVAÇÃO": "Nenhuma sobra identificada."}]

            gerar_pdf_tabela(
                "RELATÓRIO DE ITENS FALTANTES",
                f"Conferência Consumo Necessário x Ordem de Compra | {di} a {df}",
                linhas_f,
                ["PRODUTO", "UN", "NECESSÁRIO", "COMPRADO", "FALTANTE", "DATA", "OBSERVAÇÃO"],
                faltantes_pdf,
            )
            gerar_pdf_tabela(
                "RELATÓRIO DE ITENS COM SOBRA",
                f"Conferência Consumo Necessário x Ordem de Compra | {di} a {df}",
                linhas_s,
                ["PRODUTO", "UN", "NECESSÁRIO", "COMPRADO", "SOBRA", "DATA", "OBSERVAÇÃO"],
                sobras_pdf,
            )
            gerar_excel(conferidos, faltantes, sobras, excel)

            self.set_status(
                "Conferência concluída com sucesso.\n\n"
                f"Itens consumo: {len(consumo)}\n"
                f"Itens OC: {len(oc)}\n"
                f"Faltantes: {max(0, len(faltantes))}\n"
                f"Sobras/excedentes: {max(0, len(sobras))}\n\n"
                f"PDF faltantes criado:\n{faltantes_pdf}\n\n"
                f"PDF sobras criado:\n{sobras_pdf}\n\n"
                f"Excel conferência criado:\n{excel}"
            )
        except Exception as e:
            self.set_status(f"ERRO na conferência:\n{e}\n\nDetalhe técnico:\n{traceback.format_exc()[-1400:]}")
        finally:
            self.busy = False


class ExpressConsumoOCApp(App):
    def build(self):
        self.title = "EXPRESS CONSUMO OC"
        return RootUI()


if __name__ == "__main__":
    ExpressConsumoOCApp().run()
